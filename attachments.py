"""공주대 공지 첨부파일/본문 이미지 → 텍스트 변환 어댑터.

각 어댑터는 실패 시 빈 문자열 또는 [실패 사유]를 돌려준다.
호출자는 결과를 본문에 그대로 이어 붙이면 된다.
"""

# --- 표준 라이브러리 ---
import io           # 바이트 데이터를 "파일처럼" 다루기 위한 BytesIO 용도 (pdfplumber/openpyxl/zipfile이 파일객체를 요구함)
import base64       # 이미지 바이트를 VLM에 보낼 때 base64 문자열로 인코딩
import logging
import os
import tempfile
import time
import zipfile      # HWPX 파일은 사실상 ZIP 컨테이너라서 직접 열어서 내부 XML을 꺼냄
from pathlib import Path                       # 파일 확장자(.pdf, .hwpx 등) 추출용
from functools import lru_cache                # VLM 클라이언트 싱글톤화에 사용
from xml.etree import ElementTree as ET        # HWPX 내부 XML 파싱

# --- 외부 라이브러리 ---
import pdfplumber                              # PDF에서 텍스트 추출 (텍스트 PDF용)
import openpyxl                                # XLSX 읽기 (현재 라우터에서는 미사용)
from langchain_core.messages import HumanMessage          # LangChain 멀티모달 메시지 포맷

from config import LLM_PROVIDER, VLM_MODEL, OPENAI_API_KEY, GOOGLE_API_KEY

logger = logging.getLogger(__name__)


# HWPX 본문(paragraph)의 XML 네임스페이스.
# 이 prefix를 붙여야 ElementTree가 <hp:t> 같은 텍스트 노드를 찾을 수 있다.
_HWPX_PARA_NS = "{http://www.hancom.co.kr/hwpml/2011/paragraph}"


def hwpx_bytes_to_text(data: bytes) -> str:
    """hwpx 패키지(ZIP)의 Contents/section*.xml에서 <hp:t> 텍스트 노드를 모두 추출.

    표/일정/문의처까지 텍스트로 들어있다. 이미지는 누락되지만 RAG 입력에는 충분.
    """
    parts = []  # 모든 section에서 모은 텍스트 조각을 담을 버퍼

    # bytes를 파일처럼 감싸서 ZipFile에 넘긴다 (디스크에 저장하지 않고 메모리에서 처리)
    with zipfile.ZipFile(io.BytesIO(data)) as z:
        # HWPX 내부 구조: Contents/section0.xml, section1.xml, ... 식으로 본문이 분할 저장돼 있음
        # 페이지 순서를 맞추기 위해 정렬해서 순회한다
        names = sorted(
            n for n in z.namelist()
            if n.startswith("Contents/section") and n.endswith(".xml")
        )
        for name in names:
            # 해당 section XML을 읽어 트리로 파싱
            root = ET.fromstring(z.read(name))
            # <hp:t> = HWPX의 "text run" 노드. 실제 글자가 담긴 leaf 노드만 골라낸다
            for t in root.iter(f"{_HWPX_PARA_NS}t"):
                if t.text:  # 빈 노드는 스킵
                    parts.append(t.text)

    # 줄바꿈으로 이어붙여 단일 문자열로 반환 (앞뒤 공백 정리)
    return "\n".join(parts).strip()


def hwp_bytes_to_text(data: bytes) -> str:
    """구버전 .hwp(OLE 컴파운드 바이너리) 텍스트 추출.

    .hwpx와 달리 ZIP이 아니라 olefile + zlib 압축 해제가 필요해서 LlamaIndex의 HWPReader에 의존.
    무거운 dep라 lazy import — 첨부에 .hwp가 실제로 나올 때만 로드.

    HWPReader.load_data는 Path를 요구해서 bytes를 임시 파일로 풀고 끝나면 unlink.
    """
    from llama_index.readers.file import HWPReader

    # delete=False + 수동 unlink: with block 안에서 reader가 다시 path로 open할 수 있게.
    tmp = tempfile.NamedTemporaryFile(suffix=".hwp", delete=False)
    try:
        tmp.write(data)
        tmp.close()
        docs = HWPReader().load_data(file=Path(tmp.name))
    finally:
        try:
            os.unlink(tmp.name)
        except OSError:
            pass
    return "\n".join((d.text or "") for d in docs).strip()


# VLM(Gemini)에게 OCR을 시킬 때 쓰는 고정 프롬프트.
# "설명 문장 붙이지 마라"가 핵심 — 안 그러면 "이 이미지는 ~에 대한 안내입니다" 같은 군더더기가 본문에 섞임.
_VLM_PROMPT = """이 이미지는 대학 공지글의 일부다. 이미지에 적힌 모든 텍스트와 표를 한국어 plain text로 빠짐없이 추출하라.
- 행사명, 일정, 신청기한, 신청방법, 문의처 등 정보 항목은 누락 없이 그대로 옮긴다.
- 표는 줄바꿈으로 항목을 구분한다.
- 장식/광고 문구도 모두 포함한다.
- 텍스트만 출력하고 설명 문장은 붙이지 않는다."""


@lru_cache(maxsize=1)  # 결과 1개 캐싱 = 사실상 싱글톤. 매번 새 클라이언트 만드는 비용을 피한다.
def _vlm():
    """VLM 클라이언트를 lazy하게 하나만 만들어 재사용한다. config.LLM_PROVIDER 토글에 따름.

    OCR은 일관성이 생명이라 temperature=0 고정.
    """
    if LLM_PROVIDER == "openai":
        from langchain_openai import ChatOpenAI
        return ChatOpenAI(model=VLM_MODEL, api_key=OPENAI_API_KEY, temperature=0)
    from langchain_google_genai import ChatGoogleGenerativeAI
    return ChatGoogleGenerativeAI(model=VLM_MODEL, google_api_key=GOOGLE_API_KEY, temperature=0)


def _image_to_text(image_bytes: bytes, mime: str) -> str:
    """이미지 바이트를 VLM에 던져 텍스트만 받아오는 저수준 헬퍼.

    provider별 image block 포맷이 달라 분기 필요:
      - OpenAI Vision: image_url 은 객체 {"url": "..."}
      - Gemini (langchain-google-genai): image_url 은 문자열 "data:..." (편의 단축형)
    """
    data_url = f"data:{mime};base64,{base64.b64encode(image_bytes).decode()}"

    if LLM_PROVIDER == "openai":
        image_block = {"type": "image_url", "image_url": {"url": data_url}}
    else:  # gemini
        image_block = {"type": "image_url", "image_url": data_url}

    msg = HumanMessage(content=[
        {"type": "text", "text": _VLM_PROMPT},
        image_block,
    ])

    resp = _vlm().invoke([msg])
    # resp.content가 가끔 list-of-blocks 형태로 올 때가 있어서 방어적으로 문자열화
    return resp.content if isinstance(resp.content, str) else str(resp.content)


def _download(url: str, context) -> bytes:
    """playwright context의 request로 가져온다 (같은 TLS·쿠키·UA).

    학교 사이트 일시 5xx/타임아웃 흡수용으로 1·2·4초 backoff 재시도.
    """
    # 일부러 requests를 안 쓰는 이유:
    #   브라우저 세션(쿠키/UA/Referer)을 그대로 들고 가야 통과되는 공지 시스템이 있다.
    #   context.request를 쓰면 별도 로그인/세션 동기화 없이 그대로 다운로드 가능.
    last_exc: Exception | None = None
    for attempt in range(3):
        try:
            resp = context.request.get(url, timeout=60000)
            if not resp.ok:
                raise RuntimeError(f"HTTP {resp.status} for {url}")
            return resp.body()
        except Exception as e:
            last_exc = e
            if attempt < 2:
                logger.warning("download 재시도 %d/2 [%s]: %s", attempt + 1, url, e)
                time.sleep(2 ** attempt)
    # 호출자(라우터)의 try/except에서 잡혀 "(처리 실패: ...)" 메시지로 변환된다
    assert last_exc is not None
    raise last_exc


def pdf_to_text(data: bytes) -> str:
    """텍스트 PDF용 1차 추출. 스캔 PDF는 빈 문자열을 반환한다."""
    with pdfplumber.open(io.BytesIO(data)) as pdf:
        # 페이지마다 텍스트를 뽑아 정리. extract_text가 None을 줄 수 있어 or ""로 방어.
        pages = [(p.extract_text() or "").strip() for p in pdf.pages]
    return "\n".join(pages).strip()


def _pdf_bytes_full(data: bytes) -> str:
    """pdfplumber 1차 → 비어있으면 pdf2image+VLM fallback."""
    body = pdf_to_text(data)
    if body:
        # 텍스트 레이어가 있는 정상 PDF: 1차 결과를 그대로 사용
        return body

    # 여기 도달했다는 건 "스캔본 PDF" = 이미지 덩어리. 페이지를 렌더링해서 OCR로 돌린다.
    # pdf2image는 무거운 의존성이라 폴백 경로에서만 lazy import.
    from pdf2image import convert_from_bytes
    images = convert_from_bytes(data, dpi=150)  # 150dpi면 OCR 품질과 속도의 합리적 절충

    chunks = []
    for im in images:
        # PIL 이미지를 PNG 바이트로 직렬화 → VLM에 전달
        buf = io.BytesIO()
        im.save(buf, format="PNG")
        chunks.append(_image_to_text(buf.getvalue(), "image/png"))
    return "\n".join(chunks).strip()


def _preview_failed(text: str) -> bool:
    """synapView 미리보기 결과가 "실패"인지 판정."""
    if not text:
        return True
    # 공주대 synapView는 변환 실패 시 페이지에 안내 문구를 그대로 박아둔다 → 텍스트로 잡힘
    if "변환이 실패" in text or "변환에 실패" in text:
        return True
    # 진짜 짧은 응답도 의심 (정상 공지면 보통 수백 자 이상)
    return len(text.strip()) < 30


# XLSX 첨부는 노이즈가 많아 기본은 임베딩 제외. 단 아래 키워드가 제목/본문에 보이면 활성화.
# 수강신청·교양·교과목 편성 같은 표 데이터가 핵심인 공지에서만 켜진다.
XLSX_KEYWORDS = ("교양", "수강신청", "교과목", "편성", "시간표", "강의계획", "개설")


def xlsx_relevant(*texts: str) -> bool:
    """제목·본문 등을 합쳐 XLSX_KEYWORDS 중 하나라도 포함하면 True."""
    blob = "\n".join(t for t in texts if t)
    return any(kw in blob for kw in XLSX_KEYWORDS)


def _detect_header_row(rows: list[list[str]]) -> int | None:
    """헤더 행 인덱스를 추정. 못 찾으면 None.

    휴리스틱: 첫 15행 중 비어있지 않은 셀이 5개 이상이고, 그 중 80% 이상이
    길이 40자 이하의 짧은 텍스트인 행 — 가장 셀 수가 많은 것을 선택.
    그 다음 행도 비슷한 셀 수를 가져야 헤더로 인정(데이터 행이 뒤따라야 함).
    """
    best_idx = None
    best_score = 0
    for i, row in enumerate(rows[:15]):
        non_empty = [c for c in row if c]
        if len(non_empty) < 5:
            continue
        short = [c for c in non_empty if len(c) <= 40]
        if len(short) < len(non_empty) * 0.8:
            continue
        # 다음 행도 충분히 채워져야 데이터 행으로 인정
        if i + 1 >= len(rows):
            continue
        next_non_empty = sum(1 for c in rows[i + 1] if c)
        if next_non_empty < len(non_empty) * 0.5:
            continue
        score = len(non_empty)
        if score > best_score:
            best_score = score
            best_idx = i
    return best_idx


def xlsx_to_text_prefixed(data: bytes) -> str:
    """XLSX 시트를 LLM 친화적 형식으로 직렬화 (full-prefix 변형).

    각 데이터 행을 `헤더1=값1 | 헤더2=값2 | ...`로 변환. 한 행만 봐도 컬럼 의미가
    명확하지만, 헤더가 행마다 반복되어 토큰 비용이 schema 방식보다 ~2배 크다.
    gpt-5-mini처럼 컨텍스트 윈도우가 큰 모델에서 비교 테스트 용도.

    헤더 탐지 실패 시 [Schema] 없이 탭 구분으로 fallback.
    """
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    out = []
    for sheet in wb.worksheets:
        out.append(f"[Sheet: {sheet.title}]")
        rows: list[list[str]] = []
        for row in sheet.iter_rows(values_only=True):
            # '~'는 단순 범위 구분자(예: 2011~2016)인데 일부 markdown 렌더러에서 strikethrough처럼
            # 해석돼 LLM/뷰어 혼동을 유발 → '-'로 정규화.
            cells = [
                "" if v is None else " ".join(str(v).split()).replace("~", "-")
                for v in row
            ]
            if any(cells):
                rows.append(cells)
        if not rows:
            continue

        header_idx = _detect_header_row(rows)
        if header_idx is None:
            for r in rows:
                out.append("\t".join(r))
            continue

        headers = list(rows[header_idx])
        for upper_idx in range(header_idx - 1, -1, -1):
            upper = rows[upper_idx]
            non_empty_count = sum(1 for c in upper if c)
            if non_empty_count <= 1:
                break
            filled = []
            cur = ""
            for c in upper:
                if c:
                    cur = c
                filled.append(cur)
            for i, c in enumerate(filled):
                if c and i < len(headers) and headers[i]:
                    headers[i] = f"{c}/{headers[i]}"

        out.append("[Headers] " + " | ".join(h for h in headers if h))

        for r in rows[header_idx + 1:]:
            parts = []
            for i, c in enumerate(r):
                if not c:
                    continue
                h = headers[i] if i < len(headers) and headers[i] else f"col{i + 1}"
                parts.append(f"{h}={c}")
            if parts:
                out.append(" | ".join(parts))
    return "\n".join(out).strip()


def xlsx_to_text(data: bytes) -> str:
    """XLSX 시트를 LLM 친화적 형식으로 직렬화 (schema 방식, 기본).

    헤더 행을 탐지해 [Schema] 라인에 컬럼명을 한 번 명시한 뒤, 데이터 행은
    탭 구분 (= 스키마 순서대로)으로 출력. LLM은 [Schema]를 보고 각 컬럼 위치를
    매핑하면 된다. row 단위 토큰 비용 최소화.

    헤더 탐지 실패 시 [Schema] 없이 모든 행을 탭 구분으로 fallback.
    """
    # data_only=True: 수식 대신 계산된 값을 가져옴
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    out = []
    for sheet in wb.worksheets:
        out.append(f"[Sheet: {sheet.title}]")
        # 1) 비어있지 않은 행만 모음 (개행/공백은 단일 공백으로 정규화 — 헤더의 줄바꿈 제거)
        #    '~'는 범위 구분자(2011~2016)인데 markdown 렌더러 일부가 strikethrough로 해석 → '-'로 정규화.
        rows: list[list[str]] = []
        for row in sheet.iter_rows(values_only=True):
            cells = [
                "" if v is None else " ".join(str(v).split()).replace("~", "-")
                for v in row
            ]
            if any(cells):
                rows.append(cells)
        if not rows:
            continue

        header_idx = _detect_header_row(rows)
        if header_idx is None:
            # 헤더 탐지 실패 — 기존 탭 fallback
            for r in rows:
                out.append("\t".join(r))
            continue

        # 2) 헤더 위쪽 행에 머지된 super-header가 있으면 prefix로 결합
        headers = list(rows[header_idx])
        for upper_idx in range(header_idx - 1, -1, -1):
            upper = rows[upper_idx]
            non_empty_count = sum(1 for c in upper if c)
            # 셀이 1개뿐(타이틀 행)이거나 모두 비면 중단
            if non_empty_count <= 1:
                break
            # 머지셀은 가장 왼쪽 셀에만 값이 있음 — forward-fill
            filled = []
            cur = ""
            for c in upper:
                if c:
                    cur = c
                filled.append(cur)
            for i, c in enumerate(filled):
                if c and i < len(headers) and headers[i]:
                    headers[i] = f"{c}/{headers[i]}"

        # 3) [Schema] 라인: 컬럼명을 1회 명시 (col1, col2, ... 인덱스 표기로 LLM이 행과 매핑)
        out.append(
            "[Schema] "
            + " | ".join(
                f"col{i + 1}={h}" for i, h in enumerate(headers) if h
            )
        )

        # 4) 데이터 행: 스키마 순서대로 탭 구분 (헤더 prefix 미포함 — 토큰 절약)
        for r in rows[header_idx + 1:]:
            out.append("\t".join(r))
    return "\n".join(out).strip()


def hwpx_via_preview(preview_url: str, context) -> str:
    """공주대 synapView.do 미리보기 페이지를 playwright로 열어 텍스트 추출.

    한글 미리보기는 iframe에 페이지를 렌더링하므로 모든 frame의 텍스트를 모은다.
    """
    page = context.new_page()
    try:
        # 네트워크 idle까지 기다린 뒤에도 2초 추가 대기 — synapView 렌더가 완전히 끝날 시간 확보
        page.goto(preview_url, wait_until="networkidle", timeout=60000)
        page.wait_for_timeout(2000)

        chunks = []

        # 1) 메인 프레임 body 텍스트 (있을 때만)
        try:
            chunks.append(page.inner_text("body"))
        except Exception as e:
            # 어떤 페이지는 main body가 비어있을 수 있음 — 그래도 iframe에서 건진다
            logger.warning("synapView main body 추출 실패 [%s]: %s", preview_url, e)

        # 2) 모든 iframe 순회 — synapView는 본문을 iframe에 렌더하는 게 일반적
        for frame in page.frames:
            if frame == page.main_frame:
                continue  # 위에서 이미 처리
            try:
                t = frame.inner_text("body")
                if t.strip():
                    chunks.append(t)
            except Exception as e:
                # 일부 frame은 cross-origin이거나 body가 없을 수 있음 → 무시하고 다음 frame
                logger.warning("synapView iframe 추출 실패 [%s, frame=%s]: %s",
                               preview_url, getattr(frame, "url", "?"), e)
                continue

        return "\n".join(chunks).strip()
    finally:
        # 예외가 나도 페이지는 반드시 닫는다 (브라우저 리소스 누수 방지)
        page.close()


def _detect_image_mime(data: bytes) -> str | None:
    """magic number로 OpenAI Vision이 받는 4개 포맷만 식별. 그 외는 None.

    OpenAI 지원: png, jpeg, gif, webp. 메일 cidimageviewer 등 동적 endpoint는
    URL에 확장자가 없어 mime 추정이 불가능 — 헤더로 판정.
    """
    if data.startswith(b"\x89PNG\r\n\x1a\n"):
        return "image/png"
    if data[:3] == b"\xff\xd8\xff":
        return "image/jpeg"
    if data[:6] in (b"GIF87a", b"GIF89a"):
        return "image/gif"
    if data[:4] == b"RIFF" and data[8:12] == b"WEBP":
        return "image/webp"
    return None


def inline_image_to_text(image_url: str, context):
    """본문 inline 이미지를 VLM으로 텍스트화.

    반환: (text, raw_bytes, mime)
      - 다운로드 실패: ("", None, None)
      - 미지원 포맷: ("", raw_bytes, None)  — bytes는 보존, VLM 호출 skip
      - 다운로드 성공/VLM 실패: ("", raw_bytes, mime)  — bytes는 보존
    """
    # 1단계: 이미지 다운로드. 실패하면 더 진행할 의미가 없으니 즉시 빈 결과 반환.
    try:
        data = _download(image_url, context)
    except Exception as e:
        logger.warning("inline image 다운로드 실패 [%s]: %s", image_url, e)
        return "", None, None

    # magic number로 실제 mime 식별. 지원 4개 포맷이 아니면 VLM skip (400 회피).
    mime = _detect_image_mime(data)
    if mime is None:
        logger.warning("inline image 미지원 포맷, VLM skip [%s]", image_url)
        return "", data, None

    # 2단계: VLM 호출. 실패해도 raw_bytes는 살려둬서 호출자가 재처리할 수 있게 한다.
    try:
        text = _image_to_text(data, mime).strip()
    except Exception as e:
        logger.warning("inline image VLM 실패 [%s]: %s", image_url, e, exc_info=True)
        text = ""

    return text, data, mime


def attachment_to_text(att: dict, context, include_xlsx: bool = False):
    """att = {'filename', 'download_url', 'preview_url' | None}.

    include_xlsx: 기본 False. True면 XLSX 본문도 추출해 임베딩 대상에 포함.
                  호출자(crawler)가 xlsx_relevant(title, body)로 판정해 전달.

    반환: (text, asset_meta)
      text: '[첨부: <파일명>]\\n<본문>' (기존 호환, 실패 시 본문 자리에 사유)
      asset_meta: {
        kind: inline_image | attachment_image | attachment_pdf
              | attachment_hwpx | attachment_xlsx | attachment_other,
        filename, source_url, mime_type,
        raw_bytes: bytes | None,  — attachment_image일 때만 채워짐
        extracted_text: str
      }
    """
    # 입력 unpack
    name = att["filename"]
    ext = Path(name.lower()).suffix       # .pdf / .hwpx / .jpg ... — 소문자 통일 후 확장자 추출
    label = f"[첨부: {name}]"             # 본문 앞에 붙일 라벨 (RAG 컨텍스트에서 출처 식별용)
    source_url = att["download_url"]

    # 메타 기본값을 먼저 깔아두고, 아래 분기에서 필드를 덮어쓰는 패턴
    meta = {
        "kind": "attachment_other",
        "filename": name,
        "source_url": source_url,
        "mime_type": None,
        "raw_bytes": None,
        "extracted_text": "",
    }

    # try 전체로 감싸서 어떤 분기에서 예외가 나더라도 "(처리 실패: ...)" 문자열로 환원한다
    # → 호출자(crawler)는 예외 처리 없이 결과를 본문에 그대로 이어붙일 수 있다
    try:
        # ───────── 분기 1: PDF ─────────
        if ext == ".pdf":
            meta["kind"] = "attachment_pdf"
            meta["mime_type"] = "application/pdf"
            data = _download(source_url, context)
            body = _pdf_bytes_full(data)   # 텍스트 1차 → 실패 시 이미지 OCR 폴백 (위 함수 참고)

        # ───────── 분기 2: 엑셀 ─────────
        elif ext in (".xlsx", ".xls"):
            meta["kind"] = "attachment_xlsx"
            meta["mime_type"] = (
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                if ext == ".xlsx" else "application/vnd.ms-excel"
            )
            if include_xlsx and ext == ".xlsx":
                # 키워드 매칭(수강신청·교양·편성 등)이 걸린 공지 → 표 전체를 텍스트화해서 임베딩 대상에 포함.
                # prefixed 방식: 행마다 `헤더=값` 명시. 토큰 비용 ~2배지만 LLM이 컬럼 의미 혼동 없이
                # 라벨 매칭만 하면 되어 추출 정확도 우월. gpt-5-mini 컨텍스트로 감당.
                data = _download(source_url, context)
                body = xlsx_to_text_prefixed(data)
            else:
                # 기본: 노이즈 많은 엑셀은 임베딩 제외하고 안내문만 남김 (.xls는 openpyxl 미지원이라 항상 제외)
                body = f"(엑셀 첨부 — 임베딩 제외. 원본 다운로드: {source_url})"

        # ───────── 분기 3: HWPX / HWP ─────────
        elif ext in (".hwpx", ".hwp"):
            meta["kind"] = "attachment_hwpx"
            meta["mime_type"] = (
                "application/vnd.hancom.hwpx" if ext == ".hwpx" else "application/x-hwp"
            )
            body = ""

            # 1차 시도: synapView 미리보기 페이지에서 텍스트 긁기
            if att.get("preview_url"):
                body = hwpx_via_preview(att["preview_url"], context)

            # 1차가 실패면 폴백 분기
            if _preview_failed(body):
                file_data = _download(source_url, context)
                if ext == ".hwpx":
                    # .hwpx는 ZIP 구조라 직접 까서 XML 텍스트 노드를 뽑을 수 있다
                    body = hwpx_bytes_to_text(file_data)
                else:
                    # .hwp(구버전 OLE 바이너리): LlamaIndex HWPReader fallback
                    body = hwp_bytes_to_text(file_data)

        # ───────── 분기 4: 이미지 첨부 ─────────
        elif ext in (".jpg", ".jpeg", ".png", ".gif"):
            meta["kind"] = "attachment_image"
            meta["mime_type"] = "image/png" if ext == ".png" else "image/jpeg"
            data = _download(source_url, context)
            meta["raw_bytes"] = data       # 멀티모달 임베딩/재처리를 위해 원본 바이트도 보존
            body = _image_to_text(data, meta["mime_type"]).strip()

        # ───────── 분기 5: 그 외 확장자 ─────────
        else:
            # 모르는 포맷은 처리 시도조차 하지 않고 안내문만 남기고 early return
            body = "(지원하지 않는 확장자, 건너뜀)"
            meta["extracted_text"] = body
            return f"{label}\n{body}", meta

    except Exception as e:
        # 모듈 docstring의 약속: 예외를 던지지 않고 "(처리 실패: ...)" 문자열로 회수
        body = f"(처리 실패: {e})"
        meta["extracted_text"] = body
        return f"{label}\n{body}", meta

    # 정상 분기(PDF / 엑셀 / HWPX 성공 / 이미지)의 공통 마무리:
    #   본문이 비어있어도 라벨 + 안내문은 보장해서 호출자가 항상 "라벨\n본문" 형태를 받게 한다
    text = f"{label}\n{body}" if body else f"{label}\n(추출 텍스트 없음)"
    meta["extracted_text"] = body
    return text, meta
